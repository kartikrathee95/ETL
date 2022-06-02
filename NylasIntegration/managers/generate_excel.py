import os
from datetime import datetime
import traceback
from openpyxl import Workbook as PyxlWorkbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import colors, Font as pyxlFont, Alignment as pyxlAlignment ,PatternFill as pyxlPatternFill
from kafka_logger import logger

def export_calendar(data, workbookname='EarningsCalendar', active=''):
    try:
        
        order={"ticker":'Ticker',
            "company":'Company',
            "date":'Date',
            "marketTime":'Market Time',
            "confirmed":'Confirmed'}
        rows = data
        if rows:
            book = PyxlWorkbook()
            sheet = book.active
            sheet.title = workbookname
            total_cols = len(data[0].keys())

            # Set filter on first row
            sheet.auto_filter.ref = 'B:' + get_column_letter(total_cols)

            blueFill = pyxlPatternFill(start_color='002060',
                                       end_color='002060',
                                       fill_type='solid')


            for i, key in enumerate(list(order.keys())):
                
                sheet.cell(1, i+2).fill = blueFill
                sheet.cell(1, i+2).value = order[key]
                sheet.cell(1, i+2).font = pyxlFont(name='Arial', bold=True, color=colors.WHITE)
                sheet.cell(1, i+2).alignment = pyxlAlignment(wrapText=True, vertical='top')
                sheet.cell(1, i+2).fill = blueFill

            for rowno, row in enumerate(rows):
                rowno = rowno + 2
                for colno, key in enumerate(order):
                    
                    if not colno:
                        sheet.cell(row=rowno, column=colno + 2).value = row.get(key,key).upper()
                    else:
                        sheet.cell(row=rowno, column=colno + 2).value = row.get(key,key)
                    sheet.cell(row=rowno, column=colno + 2).font = pyxlFont(name='Arial', bold=True)

            sheet.cell(1, 1).value = 'Date produced'
            sheet.cell(1, 1).font = pyxlFont(name='Arial', bold=True)
            sheet.cell(2, 1).value = datetime.now().date()
            sheet.cell(2, 1).alignment = pyxlAlignment(horizontal='left')

            sheet.column_dimensions[get_column_letter(1)].width = 15
            sheet.column_dimensions[get_column_letter(2)].width = 15
            sheet.column_dimensions[get_column_letter(3)].width = 15
            sheet.column_dimensions[get_column_letter(4)].width = 15
            sheet.column_dimensions[get_column_letter(5)].width = 15
            sheet.column_dimensions[get_column_letter(6)].width = 15
        
        today = datetime.today()
        today_str = str(today.year) + str(today.strftime('%m')) + str(today.strftime('%d'))
        filename = 'sentieo_' + workbookname.lower() + '_' + today_str
        filename = filename + '.xlsx'
        filepath = os.path.join(os.curdir+'/', filename)
        book.save(filepath)
        return filename, filepath
    except Exception as e:
        logger.error("Exception occured in export_calendar".format(str(traceback.format_exc())))
        return None, traceback.format_exc()


